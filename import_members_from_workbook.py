from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

if __package__:
  from .db import SessionLocal, initialize_database
  from .models import PaymentHistory, User, UserRole, utcnow
  from .security import hash_password
else:
  from db import SessionLocal, initialize_database
  from models import PaymentHistory, User, UserRole, utcnow
  from security import hash_password


IMPORT_EMAIL_DOMAIN = 'import.caperclub.local'
DEFAULT_PASSWORD = 'Caper@123'
DEFAULT_PAYMENT_MODE = 'Cash'
MAX_REASONABLE_YEAR = 2100
DEFAULT_SHEET_SPORTS = {
  'SWIMMING 2026': 'Swimming',
  'CRICKET NIRANJAN': 'Cricket',
  'ZUMBA': 'Zumba',
  'Sheet1': 'Zumba',
}
SUPPORTED_SHEETS_IN_ORDER = tuple(DEFAULT_SHEET_SPORTS.keys())
PLAN_DAYS = {
  'Monthly': 30,
  '2 Month': 60,
  'Quarterly': 90,
  'Half-Yearly': 180,
  'Yearly': 365,
}
CANONICAL_LEVEL_SUFFIXES = {
  'Monthly': 'Month',
  '2 Month': '2 Month',
  'Quarterly': 'Qty',
  'Half-Yearly': 'HY',
  'Yearly': 'Y',
}


@dataclass(slots=True)
class WorkbookRow:
  sheet_name: str
  sport: str
  excel_row: int
  member_id: str
  name: str
  mobile_number: str | None
  membership_level: str
  membership_plan: str
  payment_amount: float
  due_amount: float
  membership_start: date | None
  membership_expiry: date | None
  reg_date: date | None
  issues: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ParsedSheet:
  sheet_name: str
  sport: str
  rows: list[WorkbookRow]
  stats: dict[str, int]


def _clean_text(value: Any) -> str:
  if value is None:
    return ''
  return str(value).strip()


def _to_float(value: Any) -> float:
  if value in (None, ''):
    return 0.0
  if isinstance(value, (int, float)):
    return round(float(value), 2)

  text = _clean_text(value).replace(',', '')
  if not text:
    return 0.0

  try:
    return round(float(text), 2)
  except ValueError:
    return 0.0


def _normalize_mobile(value: Any) -> str | None:
  if isinstance(value, int):
    digits = str(value)
    return digits or None

  if isinstance(value, float):
    if value.is_integer():
      digits = str(int(value))
      return digits or None

  digits = ''.join(character for character in _clean_text(value) if character.isdigit())
  return digits or None


def _slugify(value: str) -> str:
  normalized = ''.join(character if character.isalnum() else '-' for character in value.upper())
  normalized = '-'.join(part for part in normalized.split('-') if part)
  return normalized


def _normalize_member_id(value: str, *, row_number: int) -> str:
  source = _clean_text(value) or f'SW-IMPORT-{row_number:04d}'
  normalized = _slugify(source)
  return normalized or f'SW-IMPORT-{row_number:04d}'


def _generated_member_id(
  *,
  sport: str,
  name: str,
  mobile_number: str | None,
  row_number: int,
) -> str:
  sport_key = _slugify(sport) or 'GENERAL'
  if mobile_number:
    return f'{sport_key}-{mobile_number}'

  name_key = _slugify(name)
  if name_key:
    return f'{sport_key}-{name_key}'

  return f'{sport_key}-IMPORT-{row_number:04d}'


def _maybe_date(value: Any) -> date | None:
  if isinstance(value, datetime):
    if 2000 <= value.year <= MAX_REASONABLE_YEAR:
      return value.date()
    return None

  if isinstance(value, date):
    if 2000 <= value.year <= MAX_REASONABLE_YEAR:
      return value
    return None

  return None


def _display_name(first_name: Any, last_name: Any) -> str:
  first = _clean_text(first_name)
  last = _clean_text(last_name)
  parts = [part for part in (first, last) if part]
  return ' '.join(parts)


def _infer_plan(level: str, start_date: date | None, expiry_date: date | None) -> str:
  normalized = re.sub(r'\s+', ' ', _clean_text(level).upper())

  if 'HALF' in normalized or normalized.endswith(' HY') or 'SWIMMING HY' in normalized:
    return 'Half-Yearly'
  if 'QTY' in normalized or 'QUARTER' in normalized or '3 MONTH' in normalized:
    return 'Quarterly'
  if '2 MONTH' in normalized or '2MONTH' in normalized:
    return '2 Month'
  if '6 MONTH' in normalized:
    return 'Half-Yearly'
  if normalized.endswith(' Y') or 'YEAR' in normalized or '12 MONTH' in normalized:
    return 'Yearly'
  if 'MOTH' in normalized or 'MONTH' in normalized:
    return 'Monthly'

  if start_date and expiry_date:
    duration = (expiry_date - start_date).days
    if duration >= 330:
      return 'Yearly'
    if duration >= 170:
      return 'Half-Yearly'
    if duration >= 75:
      return 'Quarterly'
    if duration >= 45:
      return '2 Month'
    if duration >= 25:
      return 'Monthly'

  return 'Custom'


def _infer_expiry(start_date: date | None, plan: str) -> date | None:
  if start_date is None:
    return None

  days = PLAN_DAYS.get(plan)
  if days is None:
    return None

  return start_date + timedelta(days=days)


def _derive_payment_status(*, due_amount: float, membership_expiry: date | None, today: date) -> str:
  if due_amount > 0:
    return 'Pending'
  if membership_expiry and membership_expiry < today:
    return 'Expired'
  return 'Paid'


def _is_import_email(email: str | None) -> bool:
  if not email:
    return False
  return email.strip().lower().endswith(f'@{IMPORT_EMAIL_DOMAIN}')


def _import_email(member_id: str) -> str:
  local_part = member_id.lower().replace('/', '-')
  return f'{local_part}@{IMPORT_EMAIL_DOMAIN}'


def _row_source(sheet_name: str, row_number: int) -> str:
  return f'Workbook import: {sheet_name} row {row_number}'


def _should_record_payment(item: WorkbookRow) -> bool:
  return bool(
    item.membership_level
    or item.payment_amount > 0
    or item.due_amount > 0
    or item.membership_expiry is not None
  )


def _normalize_header(value: Any) -> str:
  return re.sub(r'[^a-z0-9]+', '', _clean_text(value).lower())


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
  headers: dict[str, int] = {}
  for index, value in enumerate(row):
    key = _normalize_header(value)
    if key and key not in headers:
      headers[key] = index
  return headers


def _cell(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> Any:
  for name in names:
    index = headers.get(name)
    if index is not None and index < len(row):
      return row[index]
  return None


def _is_blankish(value: Any) -> bool:
  text = _clean_text(value).upper()
  return text in {'', '#N/A', '#DIV/0!', '#REF!', 'N/A'}


def _default_level_for_sport(sport: str, plan: str) -> str:
  suffix = CANONICAL_LEVEL_SUFFIXES.get(plan)
  if suffix is None:
    return ''
  return f'{sport.strip().title()} {suffix}'.strip()


def _detect_sheet_layout(worksheet: Any) -> tuple[int, dict[str, int], str]:
  max_scan_row = min(15, worksheet.max_row or 15)
  for header_row in range(1, max_scan_row + 1):
    row = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = _header_map(row)
    if not headers:
      continue

    has_expiry = 'expiration' in headers or 'expiry' in headers
    if {'regno', 'level', 'feespaid', 'start'}.issubset(headers) and has_expiry:
      return header_row, headers, 'registry'

    if {'level', 'feespaid', 'start'}.issubset(headers) and has_expiry:
      if 'lastname' in headers or 'name' in headers:
        return header_row, headers, 'standard'

    if {'startdate', 'name'}.issubset(headers) and has_expiry:
      return header_row, headers, 'compact'

  raise ValueError(f'Unsupported worksheet layout: {worksheet.title}')


def _build_workbook_row(
  *,
  sheet_name: str,
  sport: str,
  excel_row: int,
  member_id: str,
  name: str,
  mobile_number: str | None,
  membership_level: str,
  total_fee: Any,
  paid_value: Any,
  due_value: Any,
  membership_start_value: Any,
  membership_expiry_value: Any,
  reg_date_value: Any = None,
  stats: defaultdict[str, int],
) -> WorkbookRow | None:
  if not name and not mobile_number:
    stats['skipped_blank_identity'] += 1
    return None

  reg_date = _maybe_date(reg_date_value)
  membership_start = _maybe_date(membership_start_value)
  membership_expiry = _maybe_date(membership_expiry_value)
  issues: list[str] = []

  if membership_start is None and reg_date is not None:
    membership_start = reg_date
    issues.append('membership_start_from_reg_date')

  payment_amount = _to_float(paid_value)
  due_amount = _to_float(due_value)
  total_fee_amount = _to_float(total_fee)
  if _is_blankish(due_value) and total_fee_amount > payment_amount:
    due_amount = round(max(total_fee_amount - payment_amount, 0), 2)
    if due_amount > 0:
      issues.append('due_amount_inferred')

  normalized_level = _clean_text(membership_level)
  membership_plan = _infer_plan(normalized_level, membership_start, membership_expiry)

  if not normalized_level:
    normalized_level = _default_level_for_sport(sport, membership_plan)
    if normalized_level:
      issues.append('membership_level_inferred')

  if membership_expiry is None:
    inferred_expiry = _infer_expiry(membership_start, membership_plan)
    if inferred_expiry is not None:
      membership_expiry = inferred_expiry
      issues.append('membership_expiry_inferred')

  if membership_start is None and membership_expiry is None:
    stats['skipped_missing_membership_window'] += 1
    return None

  if issues:
    for issue in issues:
      stats[issue] += 1

  return WorkbookRow(
    sheet_name=sheet_name,
    sport=sport,
    excel_row=excel_row,
    member_id=member_id,
    name=name,
    mobile_number=mobile_number,
    membership_level=normalized_level,
    membership_plan=membership_plan,
    payment_amount=payment_amount,
    due_amount=due_amount,
    membership_start=membership_start,
    membership_expiry=membership_expiry,
    reg_date=reg_date,
    issues=issues,
  )


def _parse_registry_sheet(
  *,
  worksheet: Any,
  sheet_name: str,
  sport: str,
  header_row: int,
  headers: dict[str, int],
) -> ParsedSheet:
  stats: defaultdict[str, int] = defaultdict(int)
  rows: list[WorkbookRow] = []

  for excel_row, row in enumerate(
    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
    start=header_row + 1,
  ):
    name = _display_name(
      _cell(row, headers, 'firstname'),
      _cell(row, headers, 'lastname'),
    )
    mobile_number = _normalize_mobile(_cell(row, headers, 'mobileno', 'phone'))
    member_id = _normalize_member_id(_clean_text(_cell(row, headers, 'regno')), row_number=excel_row)

    item = _build_workbook_row(
      sheet_name=sheet_name,
      sport=sport,
      excel_row=excel_row,
      member_id=member_id,
      name=name,
      mobile_number=mobile_number,
      membership_level=_cell(row, headers, 'level'),
      total_fee=_cell(row, headers, 'membershipfees'),
      paid_value=_cell(row, headers, 'feespaid'),
      due_value=_cell(row, headers, 'feesdue'),
      membership_start_value=_cell(row, headers, 'start'),
      membership_expiry_value=_cell(row, headers, 'expiration', 'expiry'),
      reg_date_value=_cell(row, headers, 'regdate'),
      stats=stats,
    )
    if item is not None:
      rows.append(item)

  return ParsedSheet(sheet_name=sheet_name, sport=sport, rows=rows, stats=dict(stats))


def _parse_standard_sheet(
  *,
  worksheet: Any,
  sheet_name: str,
  sport: str,
  header_row: int,
  headers: dict[str, int],
) -> ParsedSheet:
  stats: defaultdict[str, int] = defaultdict(int)
  rows: list[WorkbookRow] = []

  for excel_row, row in enumerate(
    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
    start=header_row + 1,
  ):
    last_name = _cell(row, headers, 'lastname', 'name')
    first_name = _cell(row, headers, 'firstname', 'address')
    name = _display_name(first_name, last_name)
    mobile_number = _normalize_mobile(_cell(row, headers, 'phone', 'mobileno', 'contactno'))
    member_id = _generated_member_id(
      sport=sport,
      name=name,
      mobile_number=mobile_number,
      row_number=excel_row,
    )

    item = _build_workbook_row(
      sheet_name=sheet_name,
      sport=sport,
      excel_row=excel_row,
      member_id=member_id,
      name=name,
      mobile_number=mobile_number,
      membership_level=_cell(row, headers, 'level', 'month'),
      total_fee=_cell(row, headers, 'membershipfees'),
      paid_value=_cell(row, headers, 'feespaid'),
      due_value=_cell(row, headers, 'feesdue'),
      membership_start_value=_cell(row, headers, 'start'),
      membership_expiry_value=_cell(row, headers, 'expiration', 'expiry'),
      stats=stats,
    )
    if item is not None:
      rows.append(item)

  return ParsedSheet(sheet_name=sheet_name, sport=sport, rows=rows, stats=dict(stats))


def _parse_compact_sheet(
  *,
  worksheet: Any,
  sheet_name: str,
  sport: str,
  header_row: int,
  headers: dict[str, int],
) -> ParsedSheet:
  stats: defaultdict[str, int] = defaultdict(int)
  rows: list[WorkbookRow] = []

  for excel_row, row in enumerate(
    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
    start=header_row + 1,
  ):
    name = _clean_text(_cell(row, headers, 'name'))
    mobile_number = _normalize_mobile(_cell(row, headers, 'contactno', 'phone', 'mobileno'))
    member_id = _generated_member_id(
      sport=sport,
      name=name,
      mobile_number=mobile_number,
      row_number=excel_row,
    )

    item = _build_workbook_row(
      sheet_name=sheet_name,
      sport=sport,
      excel_row=excel_row,
      member_id=member_id,
      name=name,
      mobile_number=mobile_number,
      membership_level=_cell(row, headers, 'month'),
      total_fee=None,
      paid_value=None,
      due_value=None,
      membership_start_value=_cell(row, headers, 'startdate', 'start'),
      membership_expiry_value=_cell(row, headers, 'expiry', 'expiration'),
      reg_date_value=_cell(row, headers, 'startdate', 'start'),
      stats=stats,
    )
    if item is not None:
      rows.append(item)

  return ParsedSheet(sheet_name=sheet_name, sport=sport, rows=rows, stats=dict(stats))


def _parse_sheet(*, worksheet: Any, sport: str) -> ParsedSheet:
  header_row, headers, layout = _detect_sheet_layout(worksheet)
  if layout == 'registry':
    return _parse_registry_sheet(
      worksheet=worksheet,
      sheet_name=worksheet.title,
      sport=sport,
      header_row=header_row,
      headers=headers,
    )
  if layout == 'standard':
    return _parse_standard_sheet(
      worksheet=worksheet,
      sheet_name=worksheet.title,
      sport=sport,
      header_row=header_row,
      headers=headers,
    )
  if layout == 'compact':
    return _parse_compact_sheet(
      worksheet=worksheet,
      sheet_name=worksheet.title,
      sport=sport,
      header_row=header_row,
      headers=headers,
    )
  raise ValueError(f'Unsupported worksheet layout: {worksheet.title}')


def _group_rows(rows: list[WorkbookRow]) -> dict[str, list[WorkbookRow]]:
  grouped: dict[str, list[WorkbookRow]] = defaultdict(list)
  for item in rows:
    grouped[item.member_id].append(item)
  return grouped


def _selected_sheet_names(args: argparse.Namespace, workbook: Any) -> list[str]:
  if args.all_sheets:
    available = set(workbook.sheetnames)
    selected = [sheet for sheet in SUPPORTED_SHEETS_IN_ORDER if sheet in available]
    if not selected:
      raise ValueError('No supported data sheets were found in the workbook.')
    return selected

  if args.sheets:
    missing = [sheet for sheet in args.sheets if sheet not in workbook.sheetnames]
    if missing:
      joined = ', '.join(missing)
      raise ValueError(f'Requested sheet(s) not found in workbook: {joined}')
    return args.sheets

  raise ValueError('Provide at least one --sheet or use --all-sheets.')


def _sheet_sport(sheet_name: str, fallback_sport: str) -> str:
  return DEFAULT_SHEET_SPORTS.get(sheet_name, fallback_sport)


def _state_sort_key(item: WorkbookRow) -> tuple[date, date, date, str, int]:
  return (
    item.membership_start or date.min,
    item.membership_expiry or date.min,
    item.reg_date or date.min,
    item.sheet_name,
    item.excel_row,
  )


def _parse_args() -> argparse.Namespace:
  parser = argparse.ArgumentParser(description='Bulk import members from workbook sheets.')
  parser.add_argument('--workbook', required=True, help='Path to the workbook (.xlsx).')
  parser.add_argument(
    '--sheet',
    action='append',
    dest='sheets',
    help='Worksheet name to import. Repeat the flag to import multiple sheets.',
  )
  parser.add_argument(
    '--all-sheets',
    action='store_true',
    help='Import every supported member-data worksheet in the workbook.',
  )
  parser.add_argument(
    '--sport',
    default='Swimming',
    help='Fallback sport label when a sheet does not have a built-in sport mapping.',
  )
  parser.add_argument(
    '--default-password',
    default=DEFAULT_PASSWORD,
    help='Default password for newly created imported users.',
  )
  parser.add_argument(
    '--summary-only',
    action='store_true',
    help='Parse and summarize the workbook without connecting to the database.',
  )
  parser.add_argument(
    '--execute',
    action='store_true',
    help='Apply the import to the configured database.',
  )
  parser.add_argument(
    '--skip-init',
    action='store_true',
    help='Skip initialize_database() and use the current schema as-is.',
  )
  return parser.parse_args()


def main() -> int:
  args = _parse_args()
  workbook_path = Path(args.workbook).expanduser().resolve()

  if not workbook_path.exists():
    raise FileNotFoundError(f'Workbook not found: {workbook_path}')

  workbook = load_workbook(workbook_path, data_only=True, read_only=True)
  try:
    selected_sheet_names = _selected_sheet_names(args, workbook)
    parsed_sheets: list[ParsedSheet] = []
    all_rows: list[WorkbookRow] = []

    for sheet_name in selected_sheet_names:
      worksheet = workbook[sheet_name]
      sport = _sheet_sport(sheet_name, args.sport)
      parsed = _parse_sheet(worksheet=worksheet, sport=sport)
      parsed_sheets.append(parsed)
      all_rows.extend(parsed.rows)
  finally:
    workbook.close()

  grouped = _group_rows(all_rows)

  print(f'Workbook: {workbook_path}')
  print(f'Sheets selected: {", ".join(selected_sheet_names)}')
  for parsed in parsed_sheets:
    distinct_members = len({item.member_id for item in parsed.rows})
    print(
      f'  - {parsed.sheet_name} [{parsed.sport}]: '
      f'{len(parsed.rows)} rows, {distinct_members} distinct members'
    )
    if parsed.stats:
      for key in sorted(parsed.stats):
        print(f'      {key}: {parsed.stats[key]}')

  print(f'Total parsed membership rows: {len(all_rows)}')
  print(f'Total distinct members: {len(grouped)}')

  if args.summary_only and not args.execute:
    return 0

  if not args.skip_init:
    initialize_database()
  db = SessionLocal()
  today = date.today()
  created = 0
  updated = 0
  payments_added = 0

  try:
    member_ids = list(grouped.keys())
    existing_users = db.scalars(select(User).where(User.member_id.in_(member_ids))).all()
    user_by_member_id = {user.member_id: user for user in existing_users}

    existing_payment_sources: dict[str, set[str]] = defaultdict(set)
    existing_user_ids = [user.id for user in existing_users]
    if existing_user_ids:
      payment_rows = db.scalars(
        select(PaymentHistory).where(PaymentHistory.user_id.in_(existing_user_ids))
      ).all()
      for payment in payment_rows:
        existing_payment_sources[payment.user_id].add(payment.source)

    for member_id, member_rows in grouped.items():
      latest = max(member_rows, key=_state_sort_key)
      user = user_by_member_id.get(member_id)
      payment_status = _derive_payment_status(
        due_amount=latest.due_amount,
        membership_expiry=latest.membership_expiry,
        today=today,
      )
      note = (
        f'Imported from workbook "{workbook_path.name}" '
        f'({latest.sheet_name}, latest row {latest.excel_row}).'
      )

      if user is None:
        user = User(
          name=latest.name,
          email=_import_email(member_id),
          password_hash=hash_password(args.default_password),
          role=UserRole.USER,
          mobile_number=latest.mobile_number,
          member_id=member_id,
          sport=latest.sport,
          membership_plan=latest.membership_plan,
          membership_level=latest.membership_level,
          membership_start=latest.membership_start,
          membership_expiry=latest.membership_expiry,
          payment_amount=latest.payment_amount,
          due_amount=latest.due_amount,
          payment_mode=DEFAULT_PAYMENT_MODE,
          payment_status=payment_status,
          face_images_count=0,
          note=note,
          updated_at=utcnow(),
        )
        db.add(user)
        db.flush()
        user_by_member_id[member_id] = user
        created += 1
      else:
        if _is_import_email(user.email):
          user.email = _import_email(member_id)
        user.name = latest.name
        user.mobile_number = latest.mobile_number
        user.sport = latest.sport
        user.membership_plan = latest.membership_plan
        user.membership_level = latest.membership_level
        user.membership_start = latest.membership_start
        user.membership_expiry = latest.membership_expiry
        user.payment_amount = latest.payment_amount
        user.due_amount = latest.due_amount
        user.payment_mode = DEFAULT_PAYMENT_MODE
        user.payment_status = payment_status
        if not user.note or user.note.startswith('Imported from workbook "'):
          user.note = note
        user.updated_at = utcnow()
        updated += 1

      for item in sorted(member_rows, key=_state_sort_key):
        if not _should_record_payment(item):
          continue

        source = _row_source(item.sheet_name, item.excel_row)
        if source in existing_payment_sources[user.id]:
          continue

        db.add(
          PaymentHistory(
            user=user,
            plan=item.membership_plan,
            amount=item.payment_amount,
            payment_mode=DEFAULT_PAYMENT_MODE,
            payment_status=_derive_payment_status(
              due_amount=item.due_amount,
              membership_expiry=item.membership_expiry,
              today=today,
            ),
            membership_start=item.membership_start,
            membership_expiry=item.membership_expiry,
            source=source,
            created_at=datetime.combine(
              item.reg_date or item.membership_start or today,
              datetime.min.time(),
            ),
          )
        )
        existing_payment_sources[user.id].add(source)
        payments_added += 1

    if not args.execute:
      db.rollback()
      print('Dry run only. No database changes were committed.')
    else:
      db.commit()
      print('Import committed.')

    print(f'Users created: {created}')
    print(f'Users updated: {updated}')
    print(f'Payment history rows added: {payments_added}')
    if created:
      print(f'Default password for newly created users: {args.default_password}')
    return 0
  finally:
    db.close()


if __name__ == '__main__':
  raise SystemExit(main())
